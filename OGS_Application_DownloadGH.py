import selenium
import time
import datetime
import glob
import os
import openpyxl
import smtplib
import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl import load_workbook


if getattr(sys, 'frozen', False) :
    # running in a bundle
    chromedriver_path = os.path.join(sys._MEIPASS, 'chromedriver')


import readline
readline.parse_and_bind("control-v: paste")

print("Enter the folder location of the OGS_MasterSpreadsheet.xlsx. Don't forget to add a forward slash at the end of the path \ .")
masterspreadsheet_folder = input()
#masterspreadsheet = str(chromedriver_path) + '\\OGS_MasterSpreadsheet.xlsx'
masterspreadsheet = str(masterspreadsheet_folder) + 'OGS_MasterSpreadsheet.xlsx'
wb = load_workbook(masterspreadsheet)
ws = wb.active

print("Please enter the folder location of where downloads on your Chrome browser end up. Don't forget to add a forward slash at the end of the path \ .")
directory = input()

Unsuccessful_Application_Downloads = []
driver = webdriver.Chrome()
wait = WebDriverWait(driver, 60)
driver.get("https://ogs.sgs.utoronto.ca/applicant/(main:submitted-forms)")
Login = driver.find_element_by_partial_link_text("Proceed to Login")
Login.click()
Login_UTORid = driver.find_element_by_partial_link_text("Log in with UTORid / JOINid")
Login_UTORid.click()
UTORid = driver.find_element_by_id("username")
Password = driver.find_element_by_id("password")

print ('Please enter your UTORid:')
UTORid_Input = input()
print ('Please enter your password:')
Password_Input = input ()
print(100* "\n")


UTORid.send_keys(UTORid_Input)
Password.send_keys(Password_Input)
Login_button = driver.find_element_by_name("_eventId_proceed")
Login_button.click()

try:
    element = WebDriverWait(driver, 120).until(
        EC.presence_of_element_located((By.ID, "cdk-accordion-child-0"))
    )
    elem = driver.find_element_by_partial_link_text("My Forms")
    elem.click()
except:
    print('could not load OGS application site')
        
try:
    element = WebDriverWait(driver, 120).until(
        EC.presence_of_element_located((By.ID, "mat-checkbox-2"))
    )
    elem = driver.find_element_by_partial_link_text("My Forms")
    elem.click()
except:
    print('could not load OGS application site')
    

i = 1 #have to define i creater than 0 at some point...sigh.
n = 0

row_num = ws.max_row + 1

for i in range(2, row_num):

    n = n + 1
    Last_Name = ws.cell(row = i, column = 4)
    First_Name = ws.cell(row = i, column = 3)
    Graduate_Unit = ws.cell(row = i, column = 2)
    FirstLast_Name = ws.cell(row = i, column = 5) #search appears to query the referee student name field
    if FirstLast_Name.value is None:
        break
    searchbox = driver.find_element_by_xpath("//*[@class='searchbox']")
    searchbox.click()
    searchbox.send_keys(Keys.CONTROL + "a")
    searchbox.send_keys(Keys.DELETE)
    searchbox.send_keys(FirstLast_Name.value.replace("&#39;", "'"))

    try:
        j = driver.find_element_by_id("mat-checkbox-1")
        j.click()
        if n > 1:
            j.click()

    except:
        Unsuccessful_Application_Downloads.append(FirstLast_Name.value)
        print(str(datetime.datetime.now().strftime("%I:%M:%S %p")) + " Download of " + str(FirstLast_Name.value) + "'s application was unsuccessful.")
        continue

    try:
        download_button = driver.find_element_by_xpath("//*[@title='Bulk Download']")
        download_button.click()
        PDF_button = driver.find_element_by_xpath("//button/mat-icon[1]")
        PDF_button.click()
        element = WebDriverWait(driver, 150).until(
            EC.invisibility_of_element_located((By.XPATH, "//*[@class='spinner-container ng-star-inserted']"))
        )
        time.sleep(1.5)
    
    except:
        EC.visibility_of_element_located((By.XPATH, "//*[@class='spinner-container ng-star-inserted']"))
        print("still loading")
        htmlElem = driver.find_element_by_tag_name('html')
        htmlElem.send_keys(Keys.F5)
        driver.get("https://ogs.sgs.utoronto.ca/applicant/(main:submitted-forms)")
        try:
            element = WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.ID, "cdk-accordion-child-0"))
            )
            elem = driver.find_element_by_partial_link_text("My Forms")
            elem.click()
        except:
            continue
        try:
            element = WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.ID, "mat-checkbox-2"))
            )
            elem = driver.find_element_by_partial_link_text("My Forms")
            elem.click()
        except:
            continue

        
        searchbox = driver.find_element_by_xpath("//*[@class='searchbox']")
        searchbox.click()
        searchbox.send_keys(FirstLast_Name.value)

        try:
            j = driver.find_element_by_id("mat-checkbox-1")
            j.click()
            download_button = driver.find_element_by_xpath("//*[@title='Bulk Download']")
            download_button.click()
            PDF_button = driver.find_element_by_xpath("//button/mat-icon[1]")
            PDF_button.click()
            element = WebDriverWait(driver, 150).until(
                EC.invisibility_of_element_located((By.XPATH, "//*[@class='spinner-container ng-star-inserted']"))
            )
            time.sleep(1.5)
        except:
            EC.visibility_of_element_located((By.XPATH, "//*[@class='spinner-container ng-star-inserted']"))
            print("loading unsuccessful")
            htmlElem = driver.find_element_by_tag_name('html')
            htmlElem.send_keys(Keys.F5)
            driver.get("https://ogs.sgs.utoronto.ca/applicant/(main:submitted-forms)")
            try:
                element = WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.ID, "cdk-accordion-child-0"))
                )
                elem = driver.find_element_by_partial_link_text("My Forms")
                elem.click()
            except:
                continue
            try:
                element = WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.ID, "mat-checkbox-2"))
                )
                elem = driver.find_element_by_partial_link_text("My Forms")
                elem.click()
            except:
                continue
 

# if no file name found need to loop it to try to download again

    #directory = 'C:\\Users\\joksimov\\Downloads\\'
    
    try:
        os.rename((directory + ('form-export.pdf')), (directory + str(Last_Name.value.rstrip()) + ", " + str(First_Name.value.rstrip()) + " - OGS 2020-21 - " + str(Graduate_Unit.value) + '.pdf'))
        print(str(datetime.datetime.now().strftime("%I:%M:%S %p")) + " Download of " + FirstLast_Name.value + " now complete!")
        
    except FileNotFoundError:
        if EC.visibility_of_element_located((By.XPATH, "//*[@class='spinner-container ng-star-inserted']")):
            download_button = driver.find_element_by_xpath("//*[@title='Bulk Download']")
            download_button.click()
            PDF_button = driver.find_element_by_xpath("//button/mat-icon[1]")
            PDF_button.click()
            time.sleep(0.1)
            try:
                element = WebDriverWait(driver, 150).until(
                    EC.invisibility_of_element_located((By.XPATH, "//*[@class='spinner-container ng-star-inserted']"))
                )
            except:
                EC.visibility_of_element_located((By.XPATH, "//*[@class='spinner-container ng-star-inserted']"))
                print("loading unsuccessful")
                htmlElem = driver.find_element_by_tag_name('html')
                htmlElem.send_keys(Keys.F5)
                driver.get("https://ogs.sgs.utoronto.ca/applicant/(main:submitted-forms)")
                try:
                    element = WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located((By.ID, "cdk-accordion-child-0"))
                    )
                    elem = driver.find_element_by_partial_link_text("My Forms")
                    elem.click()
                except:
                    continue
                try:
                    element = WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located((By.ID, "mat-checkbox-2"))
                    )
                    elem = driver.find_element_by_partial_link_text("My Forms")
                    elem.click()
                except:
                    continue
            finally:
                time.sleep(1.5)
            try:
                os.rename((directory + ('form-export.pdf')), (directory + str(Last_Name.value.rstrip()) + ", " + str(First_Name.value.rstrip()) + " - OGS 2020-21 - " + str(Graduate_Unit.value) + '.pdf'))
                print(str(datetime.datetime.now().strftime("%I:%M:%S %p")) + " Download of " + FirstLast_Name.value + " now complete!")
            except FileNotFoundError:
                Unsuccessful_Application_Downloads.append(FirstLast_Name.value)
                print(str(datetime.datetime.now().strftime("%I:%M:%S %p")) + " Download of " + FirstLast_Name.value + "'s application was unsuccessful.")
                continue
    except FileExistsError:
        print('The file name already exists')
    except:
        Unsuccessful_Application_Downloads.append(FirstLast_Name.value)
        print(str(datetime.datetime.now().strftime("%I:%M:%S %p")) + " Download of " + FirstLast_Name.value + "'s application was unsuccessful.")
        time.sleep(0.1)
        
    
print(str(datetime.datetime.now().strftime("%I:%M:%S %p")) + " A total of " + str(n - (len(Unsuccessful_Application_Downloads))) + " student applications have now been successfullly downloaded!")
print ("The following " + (str(len(Unsuccessful_Application_Downloads))) + " applications were unsuccessfully downloaded:")
print(Unsuccessful_Application_Downloads)

